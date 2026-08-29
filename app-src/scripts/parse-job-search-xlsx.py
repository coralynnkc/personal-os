#!/usr/bin/env python3
"""
Stage 1 of the one-time job_search_2027.xlsx import.

Reads both sheets and writes a reviewable JSON file. It writes nothing to the
database — stage 2 (import-job-search.mjs) does that. Splitting it in two is
deliberate: the tracker's Company column is free text that has to be fuzzy
matched against the research sheet's company names, and you want to eyeball
those matches (and the misses) before any rows land in Supabase.

Usage:
    python3 scripts/parse-job-search-xlsx.py [--xlsx PATH] [--out PATH]

Requires openpyxl (ships with most Python installs; else `pip install openpyxl`).
"""

import argparse
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime

DEFAULT_XLSX = os.path.expanduser("~/Documents/1-school/_toplevel/job_search_2027.xlsx")
DEFAULT_OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "job-search-import.json")

# The spreadsheet describes the 2027 new-grad cycle; every bare "Jul 10" in it
# was written during 2026.
ASSUMED_YEAR = 2026

# Sheet 1 columns, in order, mapped to the metadata keys the /jobs tab reads.
RESEARCH_COLUMNS = [
    "position_title", "company_raw", "application_opens", "typical_deadline",
    "competitiveness", "industry", "role_category", "what_youd_do",
    "technical_skills", "non_technical_skills", "salary", "interview_format",
    "why_it_fits", "notes", "apply_url", "portal_last_checked",
]

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)}
MONTHS.update({m[:3].lower(): i for m, i in list(MONTHS.items())})


def clean(v):
    """Normalise a cell to a trimmed string, or None."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s or None


def norm(name):
    """Fold a company name for matching: lowercase, no punctuation, single spaces."""
    if not name:
        return ""
    s = unicodedata.normalize("NFKD", str(name)).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def split_company(raw):
    """
    Sheet 1's company column is "Name / Sector" — but slashes also appear inside
    real names ("Citadel / Citadel Securities") and inside parentheses
    ("Block (Square / Cash App)"). Split on the first top-level " / " only, and
    never inside brackets.
    """
    if not raw:
        return None, None
    depth = 0
    for i, ch in enumerate(raw):
        if ch in "([":
            depth += 1
        elif ch in ")]":
            depth -= 1
        elif ch == "/" and depth == 0 and i > 0 and raw[i - 1] == " " and raw[i + 1:i + 2] == " ":
            return raw[:i].strip(), raw[i + 1:].strip()
    return raw.strip(), None


def parse_loose_date(text):
    """'July 10, 2026' / 'Jul 10' → ISO date. Returns None when there's no date."""
    if not text:
        return None
    m = re.search(r"([A-Za-z]{3,9})\.?\s+(\d{1,2})(?:,?\s*(\d{4}))?", str(text))
    if not m:
        return None
    month = MONTHS.get(m.group(1).lower())
    if not month:
        return None
    day = int(m.group(2))
    year = int(m.group(3)) if m.group(3) else ASSUMED_YEAR
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def competitiveness_band(text):
    """
    'Moderate — airline tech is underrated and undersubscribed vs. pure tech'
    is a sentence with a filterable word at the front. Keep the word.
    """
    if not text:
        return None
    head = re.split(r"[—–\-(,:]", str(text), maxsplit=1)[0].strip()
    return head if 0 < len(head) <= 24 else None


# Tracker Status prose → the enum in 0002_job_search.sql. Ordered: first match wins.
STATUS_RULES = [
    (r"^applied", "applied"),
    (r"^rejected", "rejected"),
    (r"^offer", "offer"),
    (r"^open", "open"),
    (r"no new grad roles", "no_roles"),
    (r"^not open", "not_open"),
    (r"listing pulled", "not_open"),
    (r"^unverified", "researching"),
    (r"under maintenance", "researching"),
]


def map_status(text):
    if not text:
        return "researching", False
    s = str(text).strip().lower()
    for pattern, status in STATUS_RULES:
        if re.search(pattern, s):
            return status, True
    return "researching", False


def match_company(tracker_name, companies):
    """
    Score tracker free text against research-sheet names. Deliberately narrow:
    exact, then prefix on a word boundary, then whole-token subset. No raw
    substring matching — that is how "AG" would end up matched to "Stage".
    Returns (index, score, reason) or (None, 0, reason).
    """
    t = norm(tracker_name)
    if not t:
        return None, 0, "empty"
    t_tokens = set(t.split())

    scored = []
    for i, c in enumerate(companies):
        n = norm(c["name"])
        if not n:
            continue
        if n == t:
            scored.append((100, i, "exact"))
        elif n.startswith(t + " "):
            scored.append((80, i, "name extends tracker text"))
        elif t.startswith(n + " "):
            scored.append((70, i, "tracker text extends name"))
        elif t_tokens and t_tokens <= set(n.split()):
            scored.append((60, i, "token subset"))

    if not scored:
        return None, 0, "no candidate"

    scored.sort(reverse=True)
    best = scored[0]
    ties = [s for s in scored if s[0] == best[0]]
    if len(ties) > 1:
        names = ", ".join(companies[i]["name"] for _, i, _ in ties)
        return None, best[0], f"ambiguous between: {names}"
    return best[1], best[0], best[2]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")

    if not os.path.exists(args.xlsx):
        sys.exit(f"Spreadsheet not found: {args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    # ── Sheet 1 → companies ────────────────────────────────────────────────
    research = wb["Job Search 2027"]
    companies = []
    by_norm = {}
    duplicates = []

    for row in list(research.iter_rows(values_only=True))[1:]:
        values = [clean(v) for v in row]
        if not any(values):
            continue
        meta = dict(zip(RESEARCH_COLUMNS, values))
        name, sector = split_company(meta.get("company_raw"))
        if not name:
            continue

        meta["competitiveness_band"] = competitiveness_band(meta.get("competitiveness"))
        meta["sector"] = sector
        # Keep the human string AND a real date, so the tab can sort on one and
        # display the other.
        meta["portal_last_checked_date"] = parse_loose_date(meta.get("portal_last_checked"))
        meta = {k: v for k, v in meta.items() if v is not None}

        key = norm(name)
        if key in by_norm:
            # "Bloomberg LP" appears twice in the sheet. Merge rather than
            # creating two entities the tracker could then match ambiguously.
            existing = companies[by_norm[key]]
            duplicates.append(name)
            for k, v in meta.items():
                existing["metadata"].setdefault(k, v)
            continue

        by_norm[key] = len(companies)
        companies.append({"name": name, "kind": "company", "metadata": meta})

    # ── Sheet 2 → applications ─────────────────────────────────────────────
    tracker = wb["Application Tracker"]
    applications = []
    unmatched = []
    unmapped_status = []

    for row in list(tracker.iter_rows(values_only=True))[1:]:
        company, applied, wave, status_text, interview, notes, outcome = (list(row) + [None] * 7)[:7]
        company = clean(company)
        if not company:
            continue

        idx, score, reason = match_company(company, companies)
        linked = companies[idx] if idx is not None else None
        meta = linked["metadata"] if linked else {}

        status, recognised = map_status(status_text)
        if not recognised and status_text:
            unmapped_status.append({"company": company, "status_text": clean(status_text)})

        # Keep the original Status prose only where it says more than the enum
        # already does. "Applied" adds nothing; "Unverified — portal nav broken"
        # does, and that detail should survive the import.
        status_source = clean(status_text)
        if status_source and norm(status_source) in {"applied", "rejected", "offer", "open"}:
            status_source = None

        # The date buried in "(checked Jul 10)" is the whole point of the
        # rebuild — pull it out into a real column. Fall back to the research
        # sheet's own Portal Last Checked when the prose carries no date.
        inline = re.search(r"\(checked ([^)]+)\)", str(status_text or ""), re.I)
        checked = parse_loose_date(inline.group(1) if inline else None) \
            or meta.get("portal_last_checked_date")

        note_text = clean(notes)
        portal_url = meta.get("apply_url")
        # Several tracker Notes cells are just the portal link.
        if note_text and re.match(r"^(https?://|[\w.-]+\.[a-z]{2,}/)", note_text):
            portal_url = note_text
            note_text = None

        record = {
            "company_name": company,
            "entity_name": linked["name"] if linked else None,
            "match": {"score": score, "reason": reason},
            "role_title": meta.get("position_title"),
            "wave": clean(wave),
            "status": status,
            "status_source": status_source,
            "portal_url": portal_url,
            "portal_last_checked": checked,
            "applied_on": clean(applied),
            "interview_on": clean(interview),
            "outcome": clean(outcome),
            "notes": note_text,
        }
        applications.append(record)
        if linked is None:
            unmatched.append({"company": company, "reason": reason})

    payload = {
        "source": args.xlsx,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "companies": companies,
        "applications": applications,
        "report": {
            "companies": len(companies),
            "merged_duplicates": duplicates,
            "applications": len(applications),
            "unmatched": unmatched,
            "unmapped_status": unmapped_status,
        },
    }

    with open(args.out, "w") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    r = payload["report"]
    print(f"Parsed {args.xlsx}")
    print(f"  companies:    {r['companies']}"
          + (f"  (merged duplicate rows: {', '.join(r['merged_duplicates'])})" if r["merged_duplicates"] else ""))
    print(f"  applications: {r['applications']}")
    print()
    print("  Company links:")
    for a in applications:
        mark = "✓" if a["entity_name"] else "·"
        target = a["entity_name"] or "— unlinked —"
        print(f"    {mark} {a['company_name']:<22} → {target:<28} [{a['match']['reason']}]")
    if r["unmatched"]:
        print()
        print(f"  {len(r['unmatched'])} unmatched (imported with entity_id = null; link by hand in the drawer):")
        for u in r["unmatched"]:
            print(f"    - {u['company']}  ({u['reason']})")
    if r["unmapped_status"]:
        print()
        print(f"  {len(r['unmapped_status'])} status cells fell through to 'researching':")
        for u in r["unmapped_status"]:
            print(f"    - {u['company']}: {u['status_text']!r}")
    print()
    print(f"Wrote {args.out}")
    print("Review it, then: node scripts/import-job-search.mjs")


if __name__ == "__main__":
    main()
